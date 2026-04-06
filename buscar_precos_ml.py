"""
Buscador de Preços - Mercado Livre
====================================
Faz 3 buscas separadas por produto:
  1. Mais vendido    → busca padrão, pega o card com MAIS VENDAS (com qtd de vendas)
  2. Menor preço     → _OrderId_PRICE
  3. Menor Full      → _Frete_Full_OrderId_PRICE

- Lê produtos e palavras negativas do Google Sheets público
- Extrai qtd de vendas do HTML dos cards (ex: "+500 vendidos")
- Salva em: precos_mercadolivre.xlsx

INSTALAÇÃO:
  python3 -m pip install playwright openpyxl requests
  python3 -m playwright install chromium

USO:
  python3 buscar_precos_ml.py
"""

import sys
import re
import json
import time
import requests
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# LINK DA SUA PLANILHA GOOGLE SHEETS (formato CSV público)
# ============================================================
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQxOYcpEBTONfs_bNLcC6r6Zo9mSJifGj7wLs5rpslugosXJhDZWyUgSj__Q6jEGYyPKZl2OV2XHamV/pub?output=csv"

# Mapa de texto de vendas para número (para ordenar)
SALES_MAP = {
    "+10000": 10000, "+5000": 5000, "+1000": 1000,
    "+500": 500, "+100": 100, "+50": 50, "+25": 25,
    "+10": 10, "+5": 5,
}


def load_products_from_sheet(url):
    print("Carregando planilha do Google Sheets...")
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        reader = csv.DictReader(io.StringIO(resp.text))
        products = []
        for row in reader:
            produto = (row.get("produto") or row.get("Produto") or list(row.values())[0]).strip()
            neg_raw = (row.get("palavras_negativas") or row.get("Palavras_negativas") or
                       (list(row.values())[1] if len(row) > 1 else "")).strip()
            if not produto:
                continue
            neg_words = [w.strip().lower() for w in neg_raw.split(",") if w.strip()] if neg_raw else []
            products.append({"name": produto, "negative": neg_words})
        print(f"  {len(products)} produtos carregados.\n")
        return products
    except Exception as e:
        print(f"ERRO ao carregar planilha: {e}")
        sys.exit(1)


def is_negative(title, negative_words):
    if not negative_words:
        return False
    return any(word in title.lower() for word in negative_words)


def parse_sales(text):
    """Extrai número de vendas de texto como '+500 vendidos'. Retorna int."""
    if not text:
        return 0
    match = re.search(r'[|\s]?\+?(\d+)\s*vendidos?', text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 0


def make_slug(query):
    slug = query.lower()
    for src, dst in [("á","a"),("à","a"),("ã","a"),("â","a"),("ä","a"),
                     ("é","e"),("è","e"),("ê","e"),("ë","e"),
                     ("í","i"),("ì","i"),("î","i"),("ï","i"),
                     ("ó","o"),("ò","o"),("õ","o"),("ô","o"),("ö","o"),
                     ("ú","u"),("ù","u"),("û","u"),("ü","u"),("ç","c")]:
        slug = slug.replace(src, dst)
    slug = re.sub(r"[^a-z0-9\s]", "", slug)
    slug = re.sub(r"\s+", "-", slug.strip())
    return slug


def build_url(query, mode="relevant"):
    slug = make_slug(query)
    if mode == "relevant":
        return f"https://lista.mercadolivre.com.br/{slug}"
    elif mode == "lowest_price":
        return f"https://lista.mercadolivre.com.br/{slug}_OrderId_PRICE_NoIndex_True"
    elif mode == "full":
        return f"https://lista.mercadolivre.com.br/{slug}_Frete_Full_OrderId_PRICE_NoIndex_True"


# JavaScript para extrair cards COM qtd de vendas do HTML
JS_EXTRACT_WITH_SALES = """
() => {
    const results = [];
    document.querySelectorAll('li.ui-search-layout__item').forEach(card => {
        try {
            const titleEl = card.querySelector('h3.poly-component__title, h3, h2');
            if (!titleEl) return;
            const title = titleEl.innerText.trim();
            if (!title) return;

            const linkEl = card.querySelector('h3 a, h2 a, a');
            if (!linkEl) return;
            let link = linkEl.href || '';
            if (!link) return;
            if (link.includes('mercadolivre.com.br') && !link.includes('click1.')) {
                link = link.split('#')[0];
            }

            const intEl = card.querySelector('span.andes-money-amount__fraction, span[class*="money-amount__fraction"]');
            if (!intEl) return;
            const decEl = card.querySelector('span.andes-money-amount__cents, span[class*="money-amount__cents"]');
            const intStr = intEl.innerText.replace(/\\D/g, '');
            const decStr = decEl ? decEl.innerText.replace(/\\D/g, '').padEnd(2,'0').slice(0,2) : '00';
            if (!intStr) return;
            const price = parseFloat(intStr + '.' + decStr);
            if (isNaN(price) || price <= 0) return;

            // Extrai texto de vendas: o ML exibe "| +100 vendidos" num span
            // Busca o span com classe poly-phrase-label que contém "vendidos"
            let salesText = '';
            let salesNum = 0;
            const salesEls = card.querySelectorAll('.poly-phrase-label, [class*="phrase-label"]');
            for (const el of salesEls) {
                const t = el.innerText || '';
                if (t.includes('vendido')) {
                    const m = t.match(/\\+?(\\d+)\\s*vendidos?/i);
                    if (m) {
                        salesNum = parseInt(m[1]);
                        salesText = '+' + m[1] + ' vendidos';
                    }
                    break;
                }
            }
            // Fallback: busca no texto completo do card
            if (!salesNum) {
                const cardText = card.innerText || '';
                const m = cardText.match(/[|\\s]\\+?(\\d+)\\s*vendidos?/i);
                if (m) {
                    salesNum = parseInt(m[1]);
                    salesText = '+' + m[1] + ' vendidos';
                }
            }

            results.push({ title, link, price, salesText, salesNum });
        } catch(e) {}
    });
    return results;
}
"""

# JavaScript para extrair do JSON @graph (mais rápido, para menor preço e Full)
def extract_items_from_json(page):
    try:
        content = page.evaluate("() => document.body.innerHTML")
        match = re.search(r'"@graph"\s*:\s*(\[.*?\])\s*\}', content, re.DOTALL)
        if not match:
            return []
        graph = json.loads(match.group(1))
        items = []
        for node in graph:
            if node.get("@type") != "Product":
                continue
            try:
                name  = node.get("name", "")
                offer = node.get("offers", {})
                price = float(offer.get("price", 0))
                url   = offer.get("url", "")
                if name and price > 0 and url:
                    url = url.split("?")[0].split("#")[0]
                    items.append({"title": name, "price": price, "link": url})
            except Exception:
                continue
        return items
    except Exception:
        return []


def fetch_most_sold(page, url, neg_words):
    """Busca o anúncio com mais vendas (ignora palavras negativas)."""
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        page.wait_for_timeout(1500)

        items = page.evaluate(JS_EXTRACT_WITH_SALES)
        if not items:
            return None

        # Filtra palavras negativas
        filtered = [i for i in items if not is_negative(i["title"], neg_words)]
        if not filtered:
            return None

        # Ordena por vendas (maior primeiro)
        filtered.sort(key=lambda x: x["salesNum"], reverse=True)
        best = filtered[0]

        return {
            "price":     best["price"],
            "title":     best["title"],
            "link":      best["link"],
            "sales_text": best["salesText"] if best["salesText"] else "—",
            "sales_num": best["salesNum"],
        }
    except Exception as e:
        return None


def fetch_cheapest(page, url, neg_words):
    """Busca o anúncio mais barato (usa JSON @graph)."""
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        page.wait_for_timeout(1500)

        items = extract_items_from_json(page)
        if not items:
            return None

        filtered = [i for i in items if not is_negative(i["title"], neg_words)]
        if not filtered:
            return None

        return min(filtered, key=lambda x: x["price"])
    except Exception:
        return None


def scrape_product(page, product):
    query     = product["name"]
    neg_words = product["negative"]

    # 1. Mais vendido (com qtd de vendas)
    rel = fetch_most_sold(page, build_url(query, "relevant"), neg_words)
    time.sleep(0.8)

    # 2. Menor preço
    low = fetch_cheapest(page, build_url(query, "lowest_price"), neg_words)
    time.sleep(0.8)

    # 3. Menor Full
    full = fetch_cheapest(page, build_url(query, "full"), neg_words)
    time.sleep(0.8)

    if not rel and not low and not full:
        return None
    return {"rel": rel, "low": low, "full": full}


def build_excel(results, filename="precos_mercadolivre.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Preços ML"

    HEADER_FILL  = PatternFill("solid", fgColor="1A5276")
    FILL_LOW     = PatternFill("solid", fgColor="1E8449")
    FILL_FULL    = PatternFill("solid", fgColor="154360")
    FILL_REL     = PatternFill("solid", fgColor="6C3483")
    ALT_FILL     = PatternFill("solid", fgColor="F2F3F4")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    NO_FULL_FILL = PatternFill("solid", fgColor="FDEDEC")

    WHITE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BOLD_FONT   = Font(name="Arial", bold=True, size=10)
    NORMAL_FONT = Font(name="Arial", size=9)
    LINK_LOW    = Font(name="Arial", size=9, color="1E8449", underline="single")
    LINK_FULL   = Font(name="Arial", size=9, color="154360", underline="single")
    LINK_REL    = Font(name="Arial", size=9, color="6C3483", underline="single")
    SALES_FONT  = Font(name="Arial", size=9, bold=True, color="6C3483")

    thin   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Linha 1: título
    ws.merge_cells("A1:K1")
    ws["A1"] = "Pesquisa de Preços — Mercado Livre"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Linha 2: grupos — agora Mais Vendido tem 4 colunas (preço, título, link, vendas)
    for rng, label, fill in [
        ("A2:A3", "Produto Buscado", HEADER_FILL),
        ("B2:D2", "Menor Preço",     FILL_LOW),
        ("E2:G2", "Menor Full",      FILL_FULL),
        ("H2:K2", "Mais Vendido",    FILL_REL),   # 4 colunas
    ]:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = label; c.font = WHITE_FONT
        c.fill = fill;   c.alignment = CENTER; c.border = BORDER

    # Linha 3: subheaders
    sub_headers = (
        [("Preço", FILL_LOW), ("Título", FILL_LOW), ("Link", FILL_LOW)] +
        [("Preço", FILL_FULL), ("Título", FILL_FULL), ("Link", FILL_FULL)] +
        [("Preço", FILL_REL), ("Título", FILL_REL), ("Link", FILL_REL), ("Vendas", FILL_REL)]
    )
    for col, (label, fill) in enumerate(sub_headers, start=2):
        c = ws.cell(row=3, column=col, value=label)
        c.font = WHITE_FONT; c.fill = fill
        c.alignment = CENTER; c.border = BORDER

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    # Dados
    for i, (product, result) in enumerate(results):
        row  = i + 4
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

        neg_str      = f"\n[-] {', '.join(product['negative'])}" if product["negative"] else ""
        prod_display = product["name"] + neg_str

        c = ws.cell(row=row, column=1, value=prod_display)
        c.font = BOLD_FONT; c.fill = fill; c.alignment = LEFT; c.border = BORDER

        if result is None:
            for col in range(2, 12):
                c = ws.cell(row=row, column=col, value="Sem resultados")
                c.font = NORMAL_FONT; c.fill = fill
                c.alignment = CENTER; c.border = BORDER
            ws.row_dimensions[row].height = 40
            continue

        def write_price_group(start_col, item, link_font, price_color, is_full_col=False):
            bg = NO_FULL_FILL if (item is None and is_full_col) else fill
            if item:
                c = ws.cell(row=row, column=start_col, value=item["price"])
                c.number_format = 'R$ #,##0.00'
                c.font = Font(name="Arial", size=9, bold=True, color=price_color)
                c.fill = bg; c.alignment = CENTER; c.border = BORDER

                c = ws.cell(row=row, column=start_col + 1, value=item["title"])
                c.font = NORMAL_FONT; c.fill = bg; c.alignment = LEFT; c.border = BORDER

                c = ws.cell(row=row, column=start_col + 2, value="Abrir anúncio")
                c.hyperlink = item["link"]
                c.font = link_font; c.fill = bg; c.alignment = CENTER; c.border = BORDER
            else:
                labels = ["Sem Full", "—", "—"] if is_full_col else ["—", "—", "—"]
                for offset, val in enumerate(labels):
                    c = ws.cell(row=row, column=start_col + offset, value=val)
                    c.font = NORMAL_FONT
                    c.fill = NO_FULL_FILL if is_full_col else fill
                    c.alignment = CENTER; c.border = BORDER

        # Menor preço (cols 2-4)
        write_price_group(2, result.get("low"),  LINK_LOW,  "1E8449")
        # Menor Full (cols 5-7)
        write_price_group(5, result.get("full"), LINK_FULL, "154360", is_full_col=True)

        # Mais vendido (cols 8-11: preço, título, link, vendas)
        rel = result.get("rel")
        if rel:
            c = ws.cell(row=row, column=8, value=rel["price"])
            c.number_format = 'R$ #,##0.00'
            c.font = Font(name="Arial", size=9, bold=True, color="6C3483")
            c.fill = fill; c.alignment = CENTER; c.border = BORDER

            c = ws.cell(row=row, column=9, value=rel["title"])
            c.font = NORMAL_FONT; c.fill = fill; c.alignment = LEFT; c.border = BORDER

            c = ws.cell(row=row, column=10, value="Abrir anúncio")
            c.hyperlink = rel["link"]
            c.font = LINK_REL; c.fill = fill; c.alignment = CENTER; c.border = BORDER

            # Coluna de vendas
            sales_val = rel.get("sales_text", "—")
            c = ws.cell(row=row, column=11, value=sales_val)
            c.font = SALES_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER
        else:
            for col, val in enumerate(["—", "—", "—", "—"], start=8):
                c = ws.cell(row=row, column=col, value=val)
                c.font = NORMAL_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER

        ws.row_dimensions[row].height = 40

    # Larguras: produto, preço, título, link, preço, título, link, preço, título, link, vendas
    for i, w in enumerate([38, 14, 46, 14, 14, 46, 14, 14, 46, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(filename)
    return filename


def main():
    products = load_products_from_sheet(GOOGLE_SHEET_URL)
    total    = len(products)

    print(f"{'='*65}")
    print(f"  Buscador de Preços - Mercado Livre")
    print(f"  Produtos: {total} | 3 buscas por produto")
    print(f"  Mais vendido (com qtd vendas) / Menor preço / Menor Full")
    print(f"{'='*65}\n")

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("ERRO: Rode: python3 -m pip install playwright && python3 -m playwright install chromium")
        sys.exit(1)

    results    = []
    errors     = 0
    full_found = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"]
        )
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
            timezone_id="America/Sao_Paulo",
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
        )

        page = context.new_page()

        print("Abrindo Mercado Livre...")
        page.goto("https://www.mercadolivre.com.br", wait_until="domcontentloaded", timeout=45000)
        page.wait_for_timeout(3000)
        print("Iniciando buscas...\n")

        for i, product in enumerate(products, start=1):
            neg_info = f" [-{len(product['negative'])} neg]" if product["negative"] else ""
            print(f"[{i:3}/{total}] {product['name'][:48]:<48}{neg_info}", end=" ", flush=True)

            result = scrape_product(page, product)

            if result is None:
                print("ERRO")
                errors += 1
            else:
                ps   = f"R${result['low']['price']:.2f}"  if result.get("low")  else "?"
                fs   = f"Full R${result['full']['price']:.2f}" if result.get("full") else "sem Full"
                rel  = result.get("rel")
                rs   = f"Vendas: {rel['sales_text']} R${rel['price']:.2f}" if rel else "?"
                print(f"OK | {ps:<12} | {fs:<18} | {rs}")
                if result.get("full"):
                    full_found += 1

            results.append((product, result))
            if i % 10 == 0:
                time.sleep(3.0)

        browser.close()

    print(f"\n{'='*65}")
    print(f"  Concluído! OK: {total-errors}/{total} | Full: {full_found} | Erros: {errors}")
    print(f"{'='*65}\n")

    print("Gerando planilha...")
    print(f"\nPlanilha salva: {build_excel(results)}\n")


if __name__ == "__main__":
    main()
